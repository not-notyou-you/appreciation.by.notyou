#!/usr/bin/env python3
"""
Script untuk generate data panitia (IMPROVED v2):
- Parse filename dari folder divisi
- Generate kode unik berdasarkan nama
- Convert foto ke WebP (<300KB)
- Generate Excel (data.xlsx) dan data.json
- IMPROVED: Detect image type by magic bytes
- IMPROVED: Incremental processing (skip existing files)
- IMPROVED: ARTHA ignore ROW prefix, all ARTHA same divisi
- IMPROVED: Better error handling
"""

import os
import sys
import json
import re
from pathlib import Path
from collections import defaultdict
from PIL import Image
import openpyxl
from openpyxl.styles import Alignment, Font

try:
    from pillow_heif import register_heif_opener
    register_heif_opener()
except ImportError:
    print("PERINGATAN: pillow-heif belum terinstall, file .HEIC akan gagal diproses.")
    print("Install dengan: pip install pillow-heif\n")

ASSETS_DIR = Path("assets")                # folder foto ASLI (input, tidak diubah/dihapus)
OUTPUT_DIR = Path("web/assets/photos")      # folder foto hasil convert WebP (output)
WEB_DATA_JSON = Path("web/data.json")       # output data.json untuk website
DIVISI_FOLDERS = [
    "00. BPH - ADHIKARA",
    "01. EVENT - SANCHARA",
    "02. DOKUM - SANCHITA",
    "03. GUARDIANS - BIRENDRA",
    "04. FNB - NAYAKA",
    "05. MEDIC - JANARDANA",
    "06. EQUIPMENT - DARAKA",
    "07. PIC - ARTHA",
    "08. PR - ANANTARA",
    "09. VISUAL - SWARNA",
    "10. WEBSITE - RACHANA",
]

SUPPORTED_FORMATS = {".jpg", ".jpeg", ".png", ".webp", ".heic", ".heif", ".bmp", ".gif"}
ANOMALIES = []
PANITIA_DATA = []
KODE_MAPPING = {}
CONVERTED_COUNT = 0  # Track jumlah file baru yang dikonvert
SKIPPED_COUNT = 0    # Track jumlah file existing yang di-skip


def detect_image_type(file_path):
    """
    Detect image type dari magic bytes.
    Return: extension (.jpg, .png, .heic, dll) atau None jika tidak recognize
    """
    try:
        with open(file_path, 'rb') as f:
            magic = f.read(12)
        
        # Check HEIF/HEIC (signature check lebih spesifik)
        if magic[4:12] == b'ftypheic' or magic[4:12] == b'ftypmif1':
            return '.heic'
        
        # Check JPEG
        if magic[:3] == b'\xff\xd8\xff':
            return '.jpg'
        
        # Check PNG
        if magic[:4] == b'\x89PNG':
            return '.png'
        
        # Check WebP
        if magic[:4] == b'RIFF' and magic[8:12] == b'WEBP':
            return '.webp'
        
        # Check BMP
        if magic[:2] == b'BM':
            return '.bmp'
        
        return None
    except Exception as e:
        return None


def extract_divisi_name(folder_name):
    """Extract divisi name dari folder name (format: 'XX. DESCRIPTION - DIVSINA')"""
    try:
        parts = folder_name.split(" - ")
        if len(parts) >= 2:
            return parts[-1].strip()
    except:
        pass
    return folder_name


def extract_nama_from_filename(filename, divisi_name=""):
    """
    Extract nama lengkap dari filename (format: 'DIVISI_NAMA.ext' atau 'DIVISI - NAMA')
    Special handling untuk ARTHA: ignore "ROW x_" prefix
    Return: (nama_lengkap, file_ext) atau (None, None) jika error
    """
    name_without_ext = Path(filename).stem
    ext = Path(filename).suffix.lower()
    
    if not name_without_ext:
        return None, None
    
    # Split by underscore atau dash
    if "_" in name_without_ext:
        parts = name_without_ext.split("_", 1)
        nama = parts[1].strip() if len(parts) > 1 else None
    elif "-" in name_without_ext:
        parts = name_without_ext.split("-", 1)
        nama = parts[1].strip() if len(parts) > 1 else None
    else:
        nama = name_without_ext
    
    # SPECIAL: ARTHA - ignore "ROW x_" prefix (semua ARTHA sama, tidak dibedakan per row)
    if divisi_name == "ARTHA" and nama:
        # Remove "ROW x_" prefix jika ada (case-insensitive)
        nama = re.sub(r'^ROW\s+\d+_\s*', '', nama, flags=re.IGNORECASE)
        nama = re.sub(r'^ROW\d+\s+', '', nama, flags=re.IGNORECASE)
    
    return nama, ext


def normalize_nama(nama):
    """Normalize nama: uppercase, remove extra spaces"""
    if not nama:
        return None
    return " ".join(nama.upper().split())


def get_words_from_nama(nama):
    """Extract words (only alpha chars) dari nama"""
    if not nama:
        return []
    words = nama.split()
    words = ["".join(c for c in word if c.isalpha()) for word in words]
    words = [w for w in words if w]
    return words


def generate_kode_optimal(nama, all_nama_list):
    """
    Generate kode OPTIMAL dari nama dengan check ALL nama list (GLOBAL):
    - 1 kata: ambil kata 1 jika GLOBAL unik
    - 2+ kata: tambah kata berikutnya sampai GLOBAL unik
    all_nama_list: list of ALL normalized names dari SEMUA divisi
    Return: kode unik atau None jika error
    """
    if not nama:
        return None
    
    words = get_words_from_nama(nama)
    if not words:
        return None
    
    # Coba dari 1 kata, 2 kata, dst sampai unik
    for num_words in range(1, len(words) + 1):
        kode = "".join(words[:num_words])
        
        # Count berapa banyak nama di all_nama_list yang JUGA punya kode ini
        # (GLOBAL check, bukan per-divisi)
        count = 0
        for other_nama in all_nama_list:
            other_words = get_words_from_nama(other_nama)
            if len(other_words) >= num_words:
                other_kode = "".join(other_words[:num_words])
                if other_kode == kode:
                    count += 1
        
        # Jika cuma 1 orang di SELURUH list yang punya kode ini → unik secara global
        if count == 1:
            return kode
    
    # Fallback: gunakan semua kata (paling aman)
    kode = "".join(words)
    return kode


def convert_to_webp(input_path, output_path, max_size_kb=300, quality=75):
    """
    Convert gambar ke WebP dengan max size
    INCREMENTAL: Jika output file sudah ada → skip (assume sudah valid)
    Return: (success: bool, file_size_kb: float atau error message, is_skipped: bool)
    """
    try:
        # INCREMENTAL CHECK: Jika output file sudah ada → skip regenerate
        if output_path.exists():
            file_size_kb = os.path.getsize(output_path) / 1024
            return True, file_size_kb, True  # success, size, is_skipped=True
        
        img = Image.open(input_path)
        
        # Jika RGBA, convert ke RGB
        if img.mode == "RGBA":
            rgb_img = Image.new("RGB", img.size, (255, 255, 255))
            rgb_img.paste(img, mask=img.split()[3] if len(img.split()) == 4 else None)
            img = rgb_img
        elif img.mode != "RGB":
            img = img.convert("RGB")
        
        # Save dengan quality, cek size
        img.save(output_path, "WEBP", quality=quality, method=6)
        
        file_size_kb = os.path.getsize(output_path) / 1024
        
        # Jika lebih dari max_size_kb, kurangi quality
        if file_size_kb > max_size_kb:
            for q in range(quality - 5, 40, -5):
                try:
                    img.save(output_path, "WEBP", quality=q, method=6)
                    file_size_kb = os.path.getsize(output_path) / 1024
                    if file_size_kb <= max_size_kb:
                        break
                except Exception as inner_err:
                    # Skip quality adjustment jika ada error
                    break
        
        return True, file_size_kb, False  # success, size, is_skipped=False
    
    except Exception as e:
        return False, str(e), False


def is_image_file(file_path):
    """
    Check apakah file adalah image berdasarkan extension atau magic bytes
    Return: (is_image: bool, extension: str)
    """
    ext = Path(file_path).suffix.lower()
    
    # Jika ada extension dan support, return True
    if ext in SUPPORTED_FORMATS:
        return True, ext
    
    # Jika tidak ada extension, coba detect dari magic bytes
    if ext == '' or ext not in SUPPORTED_FORMATS:
        detected_ext = detect_image_type(file_path)
        if detected_ext in SUPPORTED_FORMATS:
            return True, detected_ext
    
    return False, ext


def scan_all_nama_in_folder(folder_path, divisi_name=""):
    """Scan semua file di folder, extract nama (untuk pre-calculate optimal kode)
    Return: list of normalized names"""
    all_nama = []
    
    if not folder_path.exists():
        return all_nama
    
    files = [f for f in folder_path.iterdir() if f.is_file()]
    
    for file_path in files:
        is_image, ext = is_image_file(file_path)
        if not is_image:
            continue
        
        nama, _ = extract_nama_from_filename(file_path.name, divisi_name)
        if not nama:
            continue
        
        nama = normalize_nama(nama)
        if nama:
            all_nama.append(nama)
    
    return all_nama


def process_divisi_folder(folder_path, divisi_name, divisi_folder_name, all_nama_list=None):
    """Process satu folder divisi. Folder input = read-only.
    Output WebP ditulis ke web/assets/photos/[divisi_folder_name]/
    INCREMENTAL: Skip file yang sudah di-convert
    all_nama_list: list semua nama di folder ini (untuk optimal kode generation)
    """
    data_in_divisi = []
    
    if not folder_path.exists():
        print(f"⚠️  Folder tidak ditemukan: {folder_path}")
        return data_in_divisi
    
    if all_nama_list is None:
        all_nama_list = scan_all_nama_in_folder(folder_path, divisi_name)
    
    # Folder output khusus divisi ini di dalam web/assets/photos
    output_divisi_dir = OUTPUT_DIR / divisi_folder_name
    output_divisi_dir.mkdir(parents=True, exist_ok=True)
    
    files = sorted([f for f in folder_path.iterdir() if f.is_file()])
    
    for file_path in files:
        # Check apakah image file (dengan magic bytes detection)
        is_image, ext = is_image_file(file_path)
        
        if not is_image:
            ANOMALIES.append(f"SKIP: {file_path.name} (format tidak support: {ext})")
            continue
        
        # Parse nama dengan special handling ARTHA
        nama, _ = extract_nama_from_filename(file_path.name, divisi_name)
        if not nama:
            ANOMALIES.append(f"SKIP: {file_path.name} (tidak bisa parse nama)")
            continue
        
        nama = normalize_nama(nama)
        if not nama:
            ANOMALIES.append(f"SKIP: {file_path.name} (nama kosong setelah normalize)")
            continue
        
        # Generate kode OPTIMAL (dengan check all_nama_list GLOBAL)
        # generate_kode_optimal sudah menjamin kode unik secara global
        kode = generate_kode_optimal(nama, all_nama_list)
        if not kode:
            ANOMALIES.append(f"SKIP: {file_path.name} (error generate kode)")
            continue
        
        KODE_MAPPING[kode] = nama
        
        # Convert ke WebP -> tulis ke folder output (BUKAN folder assets asli)
        output_filename = f"{kode}.webp"
        output_path = output_divisi_dir / output_filename
        
        # Convert file
        success, result, is_skipped = convert_to_webp(file_path, output_path)
        if not success:
            ANOMALIES.append(f"CONVERT FAIL: {file_path.name} → {output_filename} ({result})")
            continue
        
        # Track counter
        global CONVERTED_COUNT, SKIPPED_COUNT
        if is_skipped:
            SKIPPED_COUNT += 1
        else:
            CONVERTED_COUNT += 1
        
        # Path relatif ini dipakai langsung oleh web (dari web/index.html)
        relative_path = f"assets/photos/{divisi_folder_name}/{output_filename}"
        data_in_divisi.append({
            "kode": kode,
            "nama": nama,
            "divisi": divisi_name,
            "divisi_folder": divisi_folder_name,
            "foto_path": relative_path,
            "size_kb": f"{result:.2f}",
        })
        
        # Print dengan marker jika skipped
        skip_marker = " (existing)" if is_skipped else ""
        print(f"✓ {kode:20} | {nama:40} | {result:6.2f}KB{skip_marker}")
    
    # Sort dalam divisi by nama
    data_in_divisi.sort(key=lambda x: x["nama"])
    return data_in_divisi


def handle_artha_subfolder(all_artha_nama=None):
    """Handle special case: ARTHA punya subfolder ROW 1-6
    SEMUA ARTHA SAMA (ignore row distinction)
    all_artha_nama: list semua nama ARTHA (untuk optimal kode generation)
    """
    artha_path = ASSETS_DIR / "07. PIC - ARTHA"
    artha_data = []
    
    if not artha_path.exists():
        return artha_data
    
    if all_artha_nama is None:
        # Pre-scan semua subfolder ARTHA
        all_artha_nama = []
        subfolders = sorted([d for d in artha_path.iterdir() if d.is_dir()])
        for subfolder in subfolders:
            all_artha_nama.extend(scan_all_nama_in_folder(subfolder, "ARTHA"))
    
    # Process semua subfolder ARTHA (tapi output semua ke folder 07. PIC - ARTHA, tidak per row)
    subfolders = sorted([d for d in artha_path.iterdir() if d.is_dir()])
    
    for subfolder in subfolders:
        print(f"\n📁 Processing ARTHA subfolder: {subfolder.name}")
        # Output semua ARTHA ke 1 folder: "07. PIC - ARTHA" (bukan per row)
        subfolder_data = process_divisi_folder(subfolder, "ARTHA", "07. PIC - ARTHA", all_artha_nama)
        artha_data.extend(subfolder_data)
    
    # Sort ARTHA data by nama
    artha_data.sort(key=lambda x: x["nama"])
    return artha_data


def copy_static_assets(assets_dir, output_dir):
    """
    Copy static asset folders dari assets/ ke web/assets/
    - assets/divisions/ → web/assets/divisions/
    - assets/reward/ → web/assets/reward/
    Return: (copied_count, folders_copied)
    """
    import shutil
    
    copied_count = 0
    folders_copied = []
    
    # Define static folders yang perlu di-copy
    static_folders = ["divisions", "reward"]
    
    for folder_name in static_folders:
        src_folder = assets_dir / folder_name
        dst_folder = output_dir / folder_name
        
        if not src_folder.exists():
            continue
        
        try:
            # Hapus folder destination jika sudah ada (untuk refresh)
            if dst_folder.exists():
                shutil.rmtree(dst_folder)
            
            # Copy folder
            shutil.copytree(src_folder, dst_folder)
            
            # Count files di folder
            file_count = sum(1 for _ in dst_folder.glob("**/*") if _.is_file())
            copied_count += file_count
            folders_copied.append(f"{folder_name}/ ({file_count} files)")
            
            print(f"✓ Copied: {folder_name}/ ({file_count} files)")
        except Exception as e:
            print(f"⚠️  Error copying {folder_name}/: {e}")
    
    return copied_count, folders_copied


def convert_folder_to_webp(input_folder, output_folder, folder_label="", max_size_kb=300, quality=75):
    """
    Convert semua gambar di folder ke WebP
    input_folder: folder dengan gambar original
    output_folder: folder output untuk WebP
    folder_label: label untuk print (e.g., "DIVISIONS", "REWARD")
    Return: (converted_count, skipped_count, failed_count)
    """
    if not input_folder.exists():
        print(f"⚠️  Folder tidak ditemukan: {input_folder}")
        return 0, 0, 0
    
    output_folder.mkdir(parents=True, exist_ok=True)
    
    converted = 0
    skipped = 0
    failed = 0
    
    files = sorted([f for f in input_folder.iterdir() if f.is_file()])
    
    for file_path in files:
        # Check apakah image file
        is_image, ext = is_image_file(file_path)
        if not is_image:
            continue
        
        # Output filename: BASENAME.webp
        output_filename = f"{file_path.stem}.webp"
        output_path = output_folder / output_filename
        
        # Check existing
        if output_path.exists():
            skipped += 1
            print(f"  ⊙ {file_path.name} (existing)")
            continue
        
        # Convert
        try:
            img = Image.open(file_path)
            
            # RGBA → RGB
            if img.mode == "RGBA":
                rgb_img = Image.new("RGB", img.size, (255, 255, 255))
                rgb_img.paste(img, mask=img.split()[3] if len(img.split()) == 4 else None)
                img = rgb_img
            elif img.mode != "RGB":
                img = img.convert("RGB")
            
            # Save WebP
            img.save(output_path, "WEBP", quality=quality, method=6)
            
            file_size_kb = os.path.getsize(output_path) / 1024
            
            # Reduce quality jika > max_size
            if file_size_kb > max_size_kb:
                for q in range(quality - 5, 40, -5):
                    try:
                        img.save(output_path, "WEBP", quality=q, method=6)
                        file_size_kb = os.path.getsize(output_path) / 1024
                        if file_size_kb <= max_size_kb:
                            break
                    except:
                        break
            
            converted += 1
            print(f"  ✓ {output_filename} ({file_size_kb:.2f}KB)")
        
        except Exception as e:
            failed += 1
            print(f"  ✗ {file_path.name} - ERROR: {str(e)[:50]}")
    
    return converted, skipped, failed


def cleanup_orphaned_photos(panitia_data, output_dir):
    """
    Cleanup: hapus foto WebP yang tidak terdata (tidak ada di PANITIA_DATA)
    panitia_data: list semua panitia yang valid (dengan kode)
    output_dir: folder web/assets/photos/
    Return: (deleted_count, deleted_files)
    """
    # Collect semua kode yang valid
    valid_kodes = {p["kode"] for p in panitia_data}
    
    deleted_count = 0
    deleted_files = []
    
    if not output_dir.exists():
        return deleted_count, deleted_files
    
    # Scan semua subfolder divisi di output_dir
    for divisi_folder in output_dir.iterdir():
        if not divisi_folder.is_dir():
            continue
        
        # Scan semua .webp file di folder divisi
        for webp_file in divisi_folder.glob("*.webp"):
            kode = webp_file.stem  # nama file tanpa extension
            
            # Jika kode tidak ada di valid_kodes → file sampah
            if kode not in valid_kodes:
                try:
                    webp_file.unlink()  # Hapus file
                    deleted_count += 1
                    deleted_files.append(f"{divisi_folder.name}/{webp_file.name}")
                except Exception as e:
                    print(f"⚠️  Error delete {webp_file.name}: {e}")
    
    return deleted_count, deleted_files


def main():
    print("=" * 100)
    print("🚀 PANITIA DATA GENERATOR (IMPROVED v2 - Incremental + ARTHA Fix)")
    print("=" * 100)
    
    if not ASSETS_DIR.exists():
        print(f"❌ Directory tidak ditemukan: {ASSETS_DIR}")
        sys.exit(1)
    
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"📂 Foto asli dibaca dari : {ASSETS_DIR.resolve()}")
    print(f"📂 Foto WebP disimpan ke : {OUTPUT_DIR.resolve()}")
    
    # PHASE 1: Pre-scan SEMUA nama dari SEMUA divisi (GLOBAL scan untuk uniqueness)
    print(f"\n{'=' * 100}")
    print("📊 PHASE 1: Pre-scanning SEMUA nama dari SEMUA divisi (GLOBAL)...")
    print(f"{'=' * 100}\n")
    
    all_nama_global = []  # Collect SEMUA nama dari SEMUA divisi
    nama_per_divisi = {}
    
    for folder_name in DIVISI_FOLDERS:
        if folder_name == "07. PIC - ARTHA":
            # ARTHA: scan semua subfolder
            artha_path = ASSETS_DIR / folder_name
            if artha_path.exists():
                all_artha_nama = []
                subfolders = sorted([d for d in artha_path.iterdir() if d.is_dir()])
                for subfolder in subfolders:
                    subfolder_nama = scan_all_nama_in_folder(subfolder, "ARTHA")
                    all_artha_nama.extend(subfolder_nama)
                    all_nama_global.extend(subfolder_nama)
                nama_per_divisi[folder_name] = all_artha_nama
                print(f"  {folder_name}: {len(all_artha_nama)} nama")
        else:
            folder_path = ASSETS_DIR / folder_name
            if folder_path.exists():
                divisi_name = extract_divisi_name(folder_name)
                all_nama = scan_all_nama_in_folder(folder_path, divisi_name)
                nama_per_divisi[folder_name] = all_nama
                all_nama_global.extend(all_nama)
                print(f"  {folder_name}: {len(all_nama)} nama")
    
    print(f"\n✓ Total nama GLOBAL: {len(all_nama_global)}")
    
    # PHASE 2: Process setiap divisi dengan optimal kode (menggunakan GLOBAL all_nama_global)
    print(f"\n{'=' * 100}")
    print("📁 PHASE 2: Processing dan convert foto (dengan GLOBAL kode uniqueness)...")
    print(f"{'=' * 100}")
    
    for folder_name in DIVISI_FOLDERS:
        if folder_name == "07. PIC - ARTHA":
            print(f"\n📁 Processing DIVISI: {folder_name} (dengan subfolder, ARTHA ignore ROW)")
            artha_data = handle_artha_subfolder(all_nama_global)
            PANITIA_DATA.extend(artha_data)
        else:
            folder_path = ASSETS_DIR / folder_name
            divisi_name = extract_divisi_name(folder_name)
            
            print(f"\n📁 Processing DIVISI: {divisi_name} ({folder_name})")
            # PENTING: gunakan all_nama_global (bukan all_nama per-divisi)
            divisi_data = process_divisi_folder(folder_path, divisi_name, folder_name, all_nama_global)
            PANITIA_DATA.extend(divisi_data)
    
    # Cleanup: hapus foto WebP yang tidak terdata (orphaned files)
    print(f"\n{'=' * 100}")
    print("🧹 CLEANUP: Menghapus foto sampah yang tidak terdata...")
    print(f"{'=' * 100}\n")
    
    deleted_count, deleted_files = cleanup_orphaned_photos(PANITIA_DATA, OUTPUT_DIR)
    if deleted_count > 0:
        print(f"🗑️  Deleted {deleted_count} orphaned photo(s):")
        for deleted_file in deleted_files:
            print(f"  - {deleted_file}")
    else:
        print(f"✓ Tidak ada foto sampah (semua foto terdata)")
    
    # BONUS: Convert folder divisions dan reward (NO Excel entry)
    print(f"\n{'=' * 100}")
    print("📁 BONUS: Convert folder divisions dan reward (non-panitia)...")
    print(f"{'=' * 100}\n")
    
    divisions_input = ASSETS_DIR / "divisions"
    divisions_output = Path("web/assets/divisions")
    if divisions_input.exists():
        print("📁 Processing DIVISIONS folder...")
        div_conv, div_skip, div_fail = convert_folder_to_webp(
            divisions_input, divisions_output, "DIVISIONS"
        )
        print(f"  Summary: {div_conv} converted, {div_skip} skipped, {div_fail} failed")
    else:
        print(f"⚠️  {divisions_input} tidak ditemukan (skip)")
    
    reward_input = ASSETS_DIR / "reward"
    reward_output = Path("web/assets/reward")
    if reward_input.exists():
        print("\n📁 Processing REWARD folder...")
        rew_conv, rew_skip, rew_fail = convert_folder_to_webp(
            reward_input, reward_output, "REWARD"
        )
        print(f"  Summary: {rew_conv} converted, {rew_skip} skipped, {rew_fail} failed")
    else:
        print(f"⚠️  {reward_input} tidak ditemukan (skip)")
    
    # Generate Excel
    print(f"\n{'=' * 100}")
    print(f"📊 Generating Excel...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Panitia"
    
    # Header
    headers = ["KODE", "NAMA LENGKAP", "DIVISI", "DIVISI FOLDER", "FOTO PATH", "SIZE KB"]
    ws.append(headers)
    
    # Style header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data rows
    for data in PANITIA_DATA:
        ws.append([
            data["kode"],
            data["nama"],
            data["divisi"],
            data["divisi_folder"],
            data["foto_path"],
            data["size_kb"],
        ])
    
    # Adjust column width
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 12
    
    # Center align data
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    wb.save("data.xlsx")
    print(f"✓ Excel saved: data.xlsx")
    
    # Generate data.json untuk website
    print(f"\n📄 Generating web/data.json...")
    
    divisi_list = sorted(set(d["divisi"] for d in PANITIA_DATA))
    
    json_output = {
        "panitia": [
            {
                "kode": d["kode"],
                "nama": d["nama"],
                "divisi": d["divisi"],
                "divisi_folder": d["divisi_folder"],
                "foto_path": d["foto_path"],
                "pesan": ""
            }
            for d in PANITIA_DATA
        ],
        "divisi_messages": {divisi: "" for divisi in divisi_list}
    }
    
    WEB_DATA_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(WEB_DATA_JSON, "w", encoding="utf-8") as f:
        json.dump(json_output, f, ensure_ascii=False, indent=2)
    
    print(f"✓ data.json saved: {WEB_DATA_JSON.resolve()}")
    print(f"  ⚠️  Kolom 'pesan' dan 'divisi_messages' masih kosong, isi manual atau via AI nanti.")
    
    # Summary
    print(f"\n{'=' * 100}")
    print(f"📈 SUMMARY")
    print(f"{'=' * 100}")
    print(f"Total panitia berhasil: {len(PANITIA_DATA)}")
    print(f"Total foto dikonvert baru: {CONVERTED_COUNT}")
    print(f"Total foto skip (existing): {SKIPPED_COUNT}")
    print(f"Total anomali/skip: {len(ANOMALIES)}")
    print(f"Total foto sampah dihapus: {deleted_count}")
    
    if ANOMALIES:
        print(f"\n⚠️  ANOMALI REPORT:")
        for anomaly in ANOMALIES:
            print(f"  - {anomaly}")
        
        # Save anomali ke file
        with open("anomalies.txt", "w", encoding="utf-8") as f:
            for anomaly in ANOMALIES:
                f.write(anomaly + "\n")
        print(f"\n✓ Anomali report saved: anomalies.txt")
    
    print(f"\n{'=' * 100}")


if __name__ == "__main__":
    main()